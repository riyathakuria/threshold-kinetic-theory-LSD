import openpyxl
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import re

# --- CONFIGURATION ---
STABLE_GENES = ["GALC", "ATG5", "GBA2", "PIK3C3", "ATG16L1", "DNAJC5", "CLN5", "ASAH1", 
                "BECN1", "ATG14", "CLN6", "STARD3", "CERS2", "DEGS1", "ATG3", "ATG12", 
                "UGCG", "MFSD8", "SMPD1", "ULK1", "GBA", "CLN3", "PIK3R4", "ATG7", "PSAP", "PPT2"]

DYNAMIC_GENES = ["CCL3", "IL2RB", "CERK", "IL7", "CCL2", "IFNG", "IL1A", "NPC2", "CCR2", 
                 "IL1B", "CCL4", "PPT1", "SMPD2", "IL6", "IL10", "NPC1", "IL18", "CCR5", 
                 "CCL5", "TPP1", "MAP1A", "IL7R", "CXCL10", "UGT8", "BNIP3", "CLN8"]

ALL_TARGETS = STABLE_GENES + DYNAMIC_GENES

# --- UPDATED REGION MAPPING (Based on your exact paste) ---
# The code looks for these substrings inside the location row.
REGION_MAP = {
    'MD_Thalamus': ['mediodorsal nucleus of thalamus'],
    'Visual_Cortex': ['primary visual cortex', 'area V1/17'],
    'Hippocampus': ['hippocampus', 'hippocampal formation'],
    'Cerebellum': ['cerebellar cortex'],  # Replacing Cerebral Cortex with this based on your paste
    'Striatum': ['striatum', 'caudate', 'putamen'] # Generic backup since exact name wasn't provided
}

def parse_age_to_weeks(age_str):
    age_str = str(age_str).lower().strip()
    try:
        val = float(re.findall(r"[\d\.]+", age_str)[0])
        if 'pcw' in age_str: return val
        if 'mos' in age_str: return 40 + (val * 4.3)
        if 'yrs' in age_str: return 40 + (val * 52)
        return -1
    except: return -1

# --- STEP 1: FILE SELECTION ---
print("Select your HUGE ABA Excel File...")
root = tk.Tk(); root.withdraw()
file_path = filedialog.askopenfilename(title="Select massive ABA Excel File")
if not file_path: exit()

# --- STEP 2: METADATA MAPPING ---
print("Scanning headers and mapping columns...")
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
sheet = wb.active
header_rows = list(sheet.iter_rows(min_row=1, max_row=5, values_only=True))

loc_row, age_row, gene_row = header_rows[0], header_rows[3], header_rows[4]
gene_col_idx = next(i for i, v in enumerate(gene_row) if v and "gene" in str(v).lower())

# Identify relevant columns
col_assignments = {region: [] for region in REGION_MAP}
found_counts = {region: 0 for region in REGION_MAP}

for idx, (loc, age) in enumerate(zip(loc_row, age_row)):
    if idx == gene_col_idx: continue
    age_w = parse_age_to_weeks(age)
    
    if 8 <= age_w <= (11 * 52 + 40): 
        loc_str = str(loc).lower()
        for region, keywords in REGION_MAP.items():
            # Check if ANY keyword matches the location string
            if any(k.lower() in loc_str for k in keywords):
                col_assignments[region].append(idx)
                found_counts[region] += 1
                break

print("\nColumns matched per region:")
for reg, count in found_counts.items():
    print(f"  {reg}: {count}")

if sum(found_counts.values()) == 0:
    print("ERROR: Zero columns matched. Please verify the region names again.")
    exit()

# --- STEP 3: DATA EXTRACTION ---
print("\nExtracting target genes...")
gene_data = {gene: {region: [] for region in REGION_MAP} for gene in ALL_TARGETS}

for row in sheet.iter_rows(min_row=6, values_only=True):
    symbol = row[gene_col_idx]
    if symbol in ALL_TARGETS:
        for region, indices in col_assignments.items():
            vals = [row[i] for i in indices if row[i] is not None and not isinstance(row[i], str)]
            gene_data[symbol][region].extend(vals)
wb.close()

# --- STEP 4: CALCULATION & EXPORT ---
print("Calculating stats and saving CSV files...")
master_list = []

for gene, regions in gene_data.items():
    classification = "Stable" if gene in STABLE_GENES else "Dynamic"
    global_vals = []
    reg_stats = {}
    
    for reg_name, vals in regions.items():
        if vals:
            mu = np.mean(vals)
            sigma = np.std(vals)
            cv = sigma / mu if mu > 0 else 0
            reg_stats[reg_name] = {'mean': mu, 'cv': cv}
            global_vals.extend(vals)
        else:
            reg_stats[reg_name] = {'mean': 0, 'cv': 0}
            
    global_mu = np.mean(global_vals) if global_vals else 0
    
    for reg_name, stats in reg_stats.items():
        enrichment = stats['mean'] / global_mu if global_mu > 0 else 0
        master_list.append({
            'Gene': gene,
            'Classification': classification,
            'Region': reg_name,
            'Regional_Mean': round(stats['mean'], 4),
            'Regional_CV': round(stats['cv'], 4),
            'Global_Mean': round(global_mu, 4),
            'Enrichment_Score': round(enrichment, 4)
        })

# Create individual CSVs
full_df = pd.DataFrame(master_list)
for region in REGION_MAP.keys():
    region_df = full_df[full_df['Region'] == region].copy()
    region_df.sort_values(by='Enrichment_Score', ascending=False, inplace=True)
    region_df.to_csv(f"{region}_Vulnerability_Analysis.csv", index=False)
    print(f"Created: {region}_Vulnerability_Analysis.csv")

print("\nAnalysis Complete.")